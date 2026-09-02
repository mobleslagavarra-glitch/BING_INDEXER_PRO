import os
from urllib.parse import urlparse

from openpyxl import load_workbook

from core.database import DB_FILE
from repositories.domain_repository import DomainRepository
from repositories.url_repository import UrlRepository


class ExcelImportService:

    def __init__(self):
        self.url_repository = UrlRepository()
        self.domain_repository = DomainRepository()

    @staticmethod
    def normalize_host(value):

        if not value:
            return ""

        value = str(value).strip()

        if not value or any(char.isspace() for char in value):
            return ""

        if "://" not in value:
            value = "https://" + value

        parsed = urlparse(value)

        if not parsed.hostname:
            return ""

        hostname = parsed.hostname.lower().rstrip(".")

        if not hostname:
            return ""

        if "." not in hostname:
            return ""

        if hostname.startswith(".") or hostname.endswith("."):
            return ""

        if ".." in hostname:
            return ""

        return hostname

    @staticmethod
    def normalize_url(value):

        if value is None:
            return ""

        value = str(value).strip()

        if not value:
            return ""

        return value

    def import_file(self, file_path):

        if not file_path:
            raise ValueError("No se ha seleccionado ningún archivo.")

        if not os.path.isfile(file_path):
            raise ValueError("El archivo seleccionado no existe.")

        extension = os.path.splitext(file_path)[1].lower()

        if extension != ".xlsx":
            raise ValueError("El archivo debe ser de tipo .xlsx.")

        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True
        )

        try:

            worksheet = workbook.active

            rows = worksheet.iter_rows(
                values_only=True
            )

            try:
                header = next(rows)
            except StopIteration:
                raise ValueError(
                    "El archivo Excel está vacío."
                )

            url_column = None

            for index, value in enumerate(header):

                if value is None:
                    continue

                if str(value).strip().lower() == "url":
                    url_column = index
                    break

            if url_column is None:
                raise ValueError(
                    "No se ha encontrado una columna llamada 'URL'."
                )

            domains = self.domain_repository.get_all()

            domains_by_host = {}

            for domain in domains:

                host = self.normalize_host(
                    domain.domain
                )

                if host:
                    domains_by_host[host] = domain

            existing_urls = {
                item.url.strip().lower()
                for item in self.url_repository.get_all()
                if item.url
            }

            imported = []
            duplicates = 0
            invalid = 0
            unknown_domains = 0

            seen = set()

            for row in rows:

                if url_column >= len(row):
                    continue

                url = self.normalize_url(
                    row[url_column]
                )

                if not url:
                    continue

                url_host = self.normalize_host(url)

                if not url_host:
                    invalid += 1
                    continue

                if url.lower() in existing_urls:
                    duplicates += 1
                    continue

                if url.lower() in seen:
                    duplicates += 1
                    continue

                domain = domains_by_host.get(
                    url_host
                )

                if domain is None:
                    unknown_domains += 1
                    continue

                if not domain.enabled:
                    unknown_domains += 1
                    continue

                seen.add(url.lower())

                imported.append(
                    (
                        domain.id,
                        url,
                        "PENDIENTE",
                        None,
                        ""
                    )
                )

            created = self.url_repository.create_many(
                imported
            )

            return {
                "read": len(imported) + duplicates + invalid + unknown_domains,
                "imported": created,
                "duplicates": duplicates,
                "invalid": invalid,
                "unknown_domains": unknown_domains
            }

        finally:
            workbook.close()
