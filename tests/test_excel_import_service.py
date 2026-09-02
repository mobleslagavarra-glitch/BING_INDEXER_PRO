from openpyxl import Workbook

from services.excel_import_service import ExcelImportService


def test_import_excel(monkeypatch, tmp_path):

    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://example.com/uno"])
    worksheet.append(["https://example.com/dos"])
    worksheet.append(["https://example.com/uno"])
    worksheet.append(["https://otro.com/tres"])

    workbook.save(file_path)

    class Domain:
        def __init__(self, id, domain, enabled=True):
            self.id = id
            self.domain = domain
            self.enabled = enabled

    class FakeDomainRepository:

        def get_all(self):
            return [
                Domain(1, "example.com", True)
            ]

    class FakeUrl:

        def __init__(self, url):
            self.url = url

    class FakeUrlRepository:

        def __init__(self):
            self.records = [
                FakeUrl("https://example.com/existente")
            ]
            self.created = []

        def get_all(self):
            return self.records

        def create_many(self, records):
            self.created.extend(records)
            return len(records)

    repository = FakeUrlRepository()

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: repository
    )

    service = ExcelImportService()

    result = service.import_file(
        str(file_path)
    )

    assert result["imported"] == 2
    assert result["duplicates"] == 1
    assert result["unknown_domains"] == 1

    assert len(repository.created) == 2

    assert repository.created[0][0] == 1
    assert repository.created[0][2] == "PENDIENTE"


def test_import_excel_requires_url_column(
    tmp_path
):

    file_path = tmp_path / "sin_url.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["Pagina"])
    worksheet.append(["https://example.com/prueba"])

    workbook.save(file_path)

    service = ExcelImportService()

    try:
        service.import_file(
            str(file_path)
        )
        assert False
    except ValueError as error:
        assert "URL" in str(error)


def test_import_excel_invalid_extension(tmp_path):

    file_path = tmp_path / "urls.txt"
    file_path.write_text(
        "https://example.com/prueba",
        encoding="utf-8"
    )

    service = ExcelImportService()

    try:
        service.import_file(
            str(file_path)
        )
        assert False
    except ValueError as error:
        assert ".xlsx" in str(error)

def test_import_excel_empty_file(tmp_path):

    file_path = tmp_path / "vacio.xlsx"

    workbook = Workbook()
    workbook.save(file_path)

    service = ExcelImportService()

    try:
        service.import_file(str(file_path))
        assert False
    except ValueError as error:
        assert "vacío" in str(error).lower()


def test_import_excel_missing_file(tmp_path):

    file_path = tmp_path / "no_existe.xlsx"

    service = ExcelImportService()

    try:
        service.import_file(str(file_path))
        assert False
    except ValueError as error:
        assert "no existe" in str(error).lower()


def test_import_excel_empty_path():

    service = ExcelImportService()

    try:
        service.import_file("")
        assert False
    except ValueError as error:
        assert "seleccionado" in str(error).lower()


def test_import_excel_disabled_domain(
    monkeypatch,
    tmp_path
):

    file_path = tmp_path / "desactivado.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://example.com/prueba"])

    workbook.save(file_path)

    class Domain:
        def __init__(self):
            self.id = 1
            self.domain = "example.com"
            self.enabled = False

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 0
    assert result["unknown_domains"] == 1

def test_import_excel_normalizes_domain_with_scheme_and_trailing_slash(
    monkeypatch,
    tmp_path
):

    file_path = tmp_path / "normalizado.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://EXAMPLE.COM/prueba"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "https://example.com/"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            self.records = records
            return len(records)

    repository = FakeUrlRepository()

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: repository
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["unknown_domains"] == 0
    assert repository.records[0][0] == 1
    assert repository.records[0][1] == "https://EXAMPLE.COM/prueba"


def test_import_excel_duplicate_is_case_insensitive(
    monkeypatch,
    tmp_path
):

    file_path = tmp_path / "duplicado.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://example.com/prueba"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrl:

        def __init__(self, url):
            self.url = url

    class FakeUrlRepository:

        def get_all(self):
            return [
                FakeUrl("https://EXAMPLE.COM/prueba")
            ]

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 0
    assert result["duplicates"] == 1


def test_import_excel_repeated_urls_in_same_file(
    monkeypatch,
    tmp_path
):

    file_path = tmp_path / "repetidas.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://example.com/uno"])
    worksheet.append(["https://example.com/uno"])
    worksheet.append(["https://example.com/uno"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["duplicates"] == 2


def test_import_excel_ignores_empty_rows(
    monkeypatch,
    tmp_path
):

    file_path = tmp_path / "vacias.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append([None])
    worksheet.append([""])
    worksheet.append(["   "])
    worksheet.append(["https://example.com/prueba"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["duplicates"] == 0
    assert result["invalid"] == 0
    assert result["unknown_domains"] == 0

def test_normalize_host_variants():
    assert ExcelImportService.normalize_host("example.com") == "example.com"
    assert ExcelImportService.normalize_host("https://example.com/") == "example.com"
    assert ExcelImportService.normalize_host("http://EXAMPLE.COM/") == "example.com"
    assert ExcelImportService.normalize_host("  https://Example.COM/path  ") == "example.com"
    assert ExcelImportService.normalize_host("example.com.") == "example.com"


def test_normalize_host_invalid_values():
    assert ExcelImportService.normalize_host(None) == ""
    assert ExcelImportService.normalize_host("") == ""
    assert ExcelImportService.normalize_host("   ") == ""
    assert ExcelImportService.normalize_host("https://") == ""


def test_normalize_url_values():
    assert ExcelImportService.normalize_url(None) == ""
    assert ExcelImportService.normalize_url("") == ""
    assert ExcelImportService.normalize_url("   ") == ""
    assert (
        ExcelImportService.normalize_url(
            "  https://example.com/test  "
        )
        == "https://example.com/test"
    )
    assert ExcelImportService.normalize_url(123) == "123"


def test_import_excel_url_column_with_spaces_and_uppercase(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["  URL  "])
    worksheet.append(["https://example.com/prueba"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["invalid"] == 0
    assert result["unknown_domains"] == 0


def test_import_excel_url_column_not_first(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["Nombre", "URL", "Comentario"])
    worksheet.append(
        [
            "Página 1",
            "https://example.com/pagina-1",
            "válida"
        ]
    )

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["invalid"] == 0
    assert result["unknown_domains"] == 0


def test_import_excel_invalid_urls(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://example.com/valida"])
    worksheet.append(["https://"])
    worksheet.append(["   "])
    worksheet.append(["not a valid url"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["invalid"] == 2
    assert result["unknown_domains"] == 0


def test_import_excel_unknown_domain(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://unknown.com/pagina"])

    workbook.save(file_path)

    class FakeDomainRepository:

        def get_all(self):
            return []

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 0
    assert result["invalid"] == 0
    assert result["unknown_domains"] == 1


def test_import_excel_skips_rows_without_url_column_value(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append([None])
    worksheet.append([])
    worksheet.append(["https://example.com/valida"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 1
    assert result["duplicates"] == 0
    assert result["invalid"] == 0
    assert result["unknown_domains"] == 0


def test_import_excel_existing_url_with_spaces_is_duplicate(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(["URL"])
    worksheet.append(["https://example.com/existente"])

    workbook.save(file_path)

    class Domain:
        id = 1
        domain = "example.com"
        enabled = True

    class ExistingUrl:
        url = "https://example.com/existente"

    class FakeDomainRepository:

        def get_all(self):
            return [Domain()]

    class FakeUrlRepository:

        def get_all(self):
            return [ExistingUrl()]

        def create_many(self, records):
            return len(records)

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 0
    assert result["duplicates"] == 1
    assert result["invalid"] == 0
    assert result["unknown_domains"] == 0


def test_import_excel_does_not_create_records_for_empty_import(
    tmp_path,
    monkeypatch
):
    file_path = tmp_path / "urls.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["URL"])
    worksheet.append(["https://unknown.com/pagina"])

    workbook.save(file_path)

    class FakeDomainRepository:

        def get_all(self):
            return []

    class FakeUrlRepository:

        def get_all(self):
            return []

        def create_many(self, records):
            assert records == []
            return 0

    monkeypatch.setattr(
        "services.excel_import_service.DomainRepository",
        lambda: FakeDomainRepository()
    )

    monkeypatch.setattr(
        "services.excel_import_service.UrlRepository",
        lambda: FakeUrlRepository()
    )

    service = ExcelImportService()

    result = service.import_file(str(file_path))

    assert result["imported"] == 0
    assert result["unknown_domains"] == 1

def test_normalize_host_rejects_invalid_hosts():
    assert ExcelImportService.normalize_host(
        "not a valid url"
    ) == ""

    assert ExcelImportService.normalize_host(
        "https://"
    ) == ""

    assert ExcelImportService.normalize_host(
        "example..com"
    ) == ""

    assert ExcelImportService.normalize_host(
        "example com"
    ) == ""
